"""
formula_integrity_guard.py

Standalone pre-deployment verification tool. Compares a freshly-generated
engagement output file against the pristine master template and asserts
that EVERY cell outside the known, legitimate input regions is byte-for-byte
identical -- same formula text, same static content, same merged ranges,
same sheet structure.

This tool's only job is to catch the single most catastrophic failure mode
in this project: a writer-service bug (or a side-effect of the LibreOffice
headless recalculation round-trip used elsewhere in this pipeline) silently
corrupting a formula cell in Balance Sheet, Income Statement, Change in
Equity, Cash Flow, or Notes 3.2-3.23 before it reaches a real client.

Legitimate input regions (exempt from strict identity checking):
  (a) "Trial Balance" sheet, columns D/E/F/G/N/O, rows 6-149 -- BUT ONLY
      for cells where the MASTER template itself does not already contain
      a formula there. If the master shows a formula in that range (e.g.
      a subtotal row), that cell is NOT exempt and is checked normally.
  (b) "Enter Details" sheet, cells detected as green-filled in the MASTER
      template AND not already formula-bearing in the master.

The exemption set is computed ONLY from the master file, never from the
output file, so a compromised output file cannot expand its own exemption
boundary.

Usage:
    python formula_integrity_guard.py <master.xlsx> <output.xlsx> [options]

Exit codes:
    0 -- PASS (no critical findings; warnings may still be printed)
    1 -- FAIL (one or more critical findings -- DO NOT ship this file)
    2 -- COULD NOT RUN (bad arguments, missing/corrupt files, etc.)
"""

from __future__ import annotations

import argparse
import hashlib
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
SHEET_TRIAL_BALANCE = "Trial Balance"
SHEET_ENTER_DETAILS = "Enter Details"

TB_INPUT_ROW_MIN, TB_INPUT_ROW_MAX = 6, 149
TB_INPUT_COLS = ["D", "E", "F", "G", "N", "O"]

GREEN_FILL_HINTS = {
    "FF92D050", "FFC6E0B4", "FF00B050", "FFA9D08E", "FFE2EFDA", "FF375623"
}

CRITICAL = "CRITICAL"
WARNING = "WARNING"


# ---------------------------------------------------------------------------
# Finding record
# ---------------------------------------------------------------------------
@dataclass
class Finding:
    severity: str
    category: str
    sheet: str
    cell: str
    message: str
    master_value: object = None
    output_value: object = None


@dataclass
class RunStats:
    sheets_compared: int = 0
    cells_scanned: int = 0
    cells_exempt: int = 0
    findings: list = field(default_factory=list)

    def add(self, finding: Finding) -> None:
        self.findings.append(finding)

    def criticals(self):
        return [f for f in self.findings if f.severity == CRITICAL]

    def warnings(self):
        return [f for f in self.findings if f.severity == WARNING]


# ---------------------------------------------------------------------------
# Formula / value helpers
# ---------------------------------------------------------------------------
def formula_text(cell) -> Optional[str]:
    """
    Returns the formula text of a cell if it is formula-bearing, else None.
    Defensively handles ArrayFormula objects (which expose a `.text`
    attribute rather than being a plain string) in addition to the normal
    "starts with =" string case, and cross-checks against cell.data_type
    where available so an unusual internal representation doesn't slip
    past as a "non-formula" cell.
    """
    val = cell.value
    if val is None:
        return None
    if hasattr(val, "text"):  # openpyxl ArrayFormula
        return str(val.text)
    if isinstance(val, str) and val.startswith("="):
        return val
    # Defensive fallback: some formula representations may not start with
    # "=" in unusual edge cases but still be flagged as formula type by
    # openpyxl's internal data_type marker.
    data_type = getattr(cell, "data_type", None)
    if data_type == "f" and isinstance(val, str):
        return val
    return None


def is_formula_cell(cell) -> bool:
    return formula_text(cell) is not None


def is_greenish(rgb: Optional[str]) -> bool:
    if not rgb or not isinstance(rgb, str) or len(rgb) != 8:
        return False
    if rgb.upper() in GREEN_FILL_HINTS:
        return True
    try:
        r, g, b = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
    except ValueError:
        return False
    return g > r + 15 and g > b + 15 and g > 100


def cell_fill_rgb(cell) -> Optional[str]:
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            return fill.fgColor.rgb
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# File loading with fatal-error handling
# ---------------------------------------------------------------------------
def load_workbook_or_die(path: str, role: str):
    p = Path(path)
    if not p.exists():
        print(f"FATAL: {role} file not found: {path}")
        sys.exit(2)
    try:
        return load_workbook(p, data_only=False)
    except (InvalidFileException, KeyError, OSError) as exc:
        print(f"FATAL: could not open {role} file '{path}' as a valid .xlsx workbook: {exc}")
        sys.exit(2)


def verify_master_checksum(master_path: str, expected_sha256: Optional[str]) -> None:
    """
    Optional extra paranoia layer: if the caller supplies a known-good
    SHA-256 hash of the master template, verify the file passed in on the
    command line actually matches it before trusting anything it says.
    This guards against the scenario where the "master" argument itself
    has been tampered with or the wrong file was passed by mistake --
    a failure mode that would otherwise make every other check in this
    tool meaningless, since it all depends on the master being ground
    truth.
    """
    if not expected_sha256:
        return
    sha256 = hashlib.sha256()
    with open(master_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            sha256.update(chunk)
    actual = sha256.hexdigest()
    if actual.lower() != expected_sha256.lower():
        print(
            f"FATAL: master template checksum mismatch.\n"
            f"  Expected SHA-256: {expected_sha256}\n"
            f"  Actual SHA-256:   {actual}\n"
            f"The file passed as the master template does not match the "
            f"known-good checksum. Refusing to proceed, since every check "
            f"in this tool assumes the master argument is ground truth."
        )
        sys.exit(2)
    print(f"[OK] Master template checksum verified: {actual}")


# ---------------------------------------------------------------------------
# Exemption set construction (derived ONLY from the master workbook)
# ---------------------------------------------------------------------------
def build_exempt_cells(wb_master) -> dict:
    """
    Returns {sheet_name: set(coord_strings)} of cells exempt from strict
    formula/content identity checking. Computed exclusively from the
    master workbook so the output file can never influence its own
    exemption boundary.
    """
    exempt: dict = {SHEET_TRIAL_BALANCE: set(), SHEET_ENTER_DETAILS: set()}

    if SHEET_TRIAL_BALANCE in wb_master.sheetnames:
        ws = wb_master[SHEET_TRIAL_BALANCE]
        for row in range(TB_INPUT_ROW_MIN, TB_INPUT_ROW_MAX + 1):
            for col in TB_INPUT_COLS:
                coord = f"{col}{row}"
                cell = ws[coord]
                if not is_formula_cell(cell):
                    exempt[SHEET_TRIAL_BALANCE].add(coord)
                # else: master already has a formula here (e.g. a subtotal
                # row happens to fall inside this range) -- deliberately
                # NOT exempted; it will be checked like any other cell.
    else:
        print(f"[WARNING] Master template has no sheet named '{SHEET_TRIAL_BALANCE}' -- "
              f"the Trial Balance exemption rule could not be applied.")

    if SHEET_ENTER_DETAILS in wb_master.sheetnames:
        ws = wb_master[SHEET_ENTER_DETAILS]
        for row in ws.iter_rows():
            for cell in row:
                rgb = cell_fill_rgb(cell)
                if is_greenish(rgb) and not is_formula_cell(cell):
                    exempt[SHEET_ENTER_DETAILS].add(cell.coordinate)
    else:
        print(f"[WARNING] Master template has no sheet named '{SHEET_ENTER_DETAILS}' -- "
              f"the Enter Details exemption rule could not be applied.")

    print(f"[INFO] Exempt cells derived from master: "
          f"{SHEET_TRIAL_BALANCE}={len(exempt[SHEET_TRIAL_BALANCE])}, "
          f"{SHEET_ENTER_DETAILS}={len(exempt[SHEET_ENTER_DETAILS])}")
    return exempt


# ---------------------------------------------------------------------------
# Structural checks: sheet names, order, visibility
# ---------------------------------------------------------------------------
def check_sheet_structure(wb_master, wb_output, stats: RunStats) -> list[str]:
    """Returns the list of sheet names common to both workbooks, in master's
    order, after recording findings for any structural discrepancies."""
    master_sheets = wb_master.sheetnames
    output_sheets = wb_output.sheetnames

    missing = [s for s in master_sheets if s not in output_sheets]
    extra = [s for s in output_sheets if s not in master_sheets]

    for s in missing:
        stats.add(Finding(
            CRITICAL, "STRUCTURAL_SHEET_MISSING", s, "-",
            f"Sheet '{s}' exists in the master template but is MISSING "
            f"from the output file entirely.",
        ))
    for s in extra:
        stats.add(Finding(
            CRITICAL, "STRUCTURAL_SHEET_EXTRA", s, "-",
            f"Sheet '{s}' exists in the output file but does NOT exist in "
            f"the master template. An output file should never contain "
            f"extra sheets.",
        ))

    common = [s for s in master_sheets if s in output_sheets]

    master_order = [s for s in master_sheets if s in common]
    output_order = [s for s in output_sheets if s in common]
    if master_order != output_order:
        stats.add(Finding(
            WARNING, "STRUCTURAL_SHEET_ORDER_CHANGED", "-", "-",
            f"Sheet order differs between master and output.\n"
            f"    Master order: {master_order}\n"
            f"    Output order: {output_order}",
        ))

    for s in common:
        m_state = wb_master[s].sheet_state
        o_state = wb_output[s].sheet_state
        if m_state != o_state:
            stats.add(Finding(
                CRITICAL, "STRUCTURAL_SHEET_VISIBILITY_CHANGED", s, "-",
                f"Sheet '{s}' visibility changed: master='{m_state}', "
                f"output='{o_state}'. A hidden calculation/workings sheet "
                f"becoming visible (or vice versa) is a structural "
                f"anomaly that should never occur from a legitimate write.",
                master_value=m_state, output_value=o_state,
            ))

    return common


# ---------------------------------------------------------------------------
# Merged cell range check
# ---------------------------------------------------------------------------
def check_merged_ranges(ws_master: Worksheet, ws_output: Worksheet,
                         sheet_name: str, stats: RunStats) -> None:
    master_ranges = {str(r) for r in ws_master.merged_cells.ranges}
    output_ranges = {str(r) for r in ws_output.merged_cells.ranges}

    removed = master_ranges - output_ranges
    added = output_ranges - master_ranges

    for r in sorted(removed):
        stats.add(Finding(
            CRITICAL, "MERGED_RANGE_REMOVED", sheet_name, r,
            f"Merged range '{r}' exists in the master template but is "
            f"MISSING in the output file. This can silently break label "
            f"alignment or, worse, split a merged formula-bearing region.",
        ))
    for r in sorted(added):
        stats.add(Finding(
            CRITICAL, "MERGED_RANGE_ADDED", sheet_name, r,
            f"Merged range '{r}' exists in the output file but does NOT "
            f"exist in the master template. Unexpected new merges can "
            f"indicate the file was re-saved by a tool that reflowed "
            f"layout (e.g. a recalculation round-trip) rather than a "
            f"clean value-only write.",
        ))


# ---------------------------------------------------------------------------
# Defined names check (bonus, on by default -- cheap and important)
# ---------------------------------------------------------------------------
def check_defined_names(wb_master, wb_output, stats: RunStats) -> None:
    """
    Workbook-level defined names can be silently referenced inside formulas
    (e.g. a named range used instead of a literal cell reference). If a
    defined name's target changes, formulas that look byte-for-byte
    identical as TEXT can nonetheless resolve to different cells. This is
    cheap to check and closes a real blind spot in a pure cell-by-cell
    formula-text comparison.
    """
    try:
        master_names = {name: defn.value for name, defn in wb_master.defined_names.items()}
    except AttributeError:
        master_names = {name: defn.attr_text for name, defn in wb_master.defined_names.items()}
    try:
        output_names = {name: defn.value for name, defn in wb_output.defined_names.items()}
    except AttributeError:
        output_names = {name: defn.attr_text for name, defn in wb_output.defined_names.items()}

    all_names = set(master_names) | set(output_names)
    for name in sorted(all_names):
        m_val = master_names.get(name, "<NOT DEFINED>")
        o_val = output_names.get(name, "<NOT DEFINED>")
        if m_val != o_val:
            stats.add(Finding(
                CRITICAL, "DEFINED_NAME_MISMATCH", "(workbook)", name,
                f"Workbook-level defined name '{name}' differs: "
                f"master target={m_val!r}, output target={o_val!r}. Any "
                f"formula referencing this name would silently resolve "
                f"differently despite looking textually identical.",
                master_value=m_val, output_value=o_val,
            ))


# ---------------------------------------------------------------------------
# Number format drift (bonus, opt-in)
# ---------------------------------------------------------------------------
def check_number_formats(ws_master: Worksheet, ws_output: Worksheet,
                          sheet_name: str, exempt_coords: set,
                          max_row: int, max_col: int, stats: RunStats) -> None:
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            coord = f"{get_column_letter(c)}{r}"
            if coord in exempt_coords:
                continue
            m_cell = ws_master[coord]
            o_cell = ws_output[coord]
            if m_cell.number_format != o_cell.number_format:
                stats.add(Finding(
                    WARNING, "NUMBER_FORMAT_DRIFT", sheet_name, coord,
                    f"Number format differs: master={m_cell.number_format!r}, "
                    f"output={o_cell.number_format!r}. Formatting drift "
                    f"(e.g. losing thousand separators or a currency "
                    f"prefix) doesn't corrupt the underlying value but "
                    f"can materially mislead a reader of the printed "
                    f"statement.",
                    master_value=m_cell.number_format,
                    output_value=o_cell.number_format,
                ))


# ---------------------------------------------------------------------------
# Core cell-by-cell comparison
# ---------------------------------------------------------------------------
def compare_sheet_cells(ws_master: Worksheet, ws_output: Worksheet,
                         sheet_name: str, exempt_coords: set,
                         stats: RunStats, flag_input_formula_as: str = CRITICAL) -> None:
    """
    Full bounding-box scan (union of both sheets' dimensions) comparing
    every cell. Deliberately brute-force rather than relying on
    iter_rows()'s notion of "used range" in only one file, since a
    corruption could manifest as extra populated rows appearing in the
    output beyond the master's original bounding box.
    """
    max_row = max(ws_master.max_row, ws_output.max_row)
    max_col = max(ws_master.max_column, ws_output.max_column)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            coord = f"{get_column_letter(c)}{r}"
            stats.cells_scanned += 1

            m_cell = ws_master[coord]
            o_cell = ws_output[coord]

            if coord in exempt_coords:
                stats.cells_exempt += 1
                # Still worth flagging if a legitimate raw-input cell
                # unexpectedly contains a formula in the output -- the
                # writer service should only ever write literal values to
                # these cells, never formulas.
                if is_formula_cell(o_cell):
                    stats.add(Finding(
                        flag_input_formula_as, "INPUT_CELL_UNEXPECTED_FORMULA",
                        sheet_name, coord,
                        f"This cell is a legitimate raw-input cell (exempt "
                        f"from strict identity checking), but the output "
                        f"file contains a FORMULA here: {formula_text(o_cell)!r}. "
                        f"The writer service should only ever write literal "
                        f"values to input cells, never formulas.",
                        master_value=m_cell.value, output_value=o_cell.value,
                    ))
                continue

            m_formula = formula_text(m_cell)
            o_formula = formula_text(o_cell)

            if (m_formula is not None) != (o_formula is not None):
                stats.add(Finding(
                    CRITICAL, "CELL_TYPE_CHANGED", sheet_name, coord,
                    f"Cell type changed between formula and non-formula. "
                    f"Master {'HAS' if m_formula else 'does NOT have'} a "
                    f"formula here; output {'HAS' if o_formula else 'does NOT have'} "
                    f"one. This is unambiguous evidence a formula cell was "
                    f"overwritten with a literal value (or vice versa).",
                    master_value=m_cell.value, output_value=o_cell.value,
                ))
                continue

            if m_formula is not None:
                if m_formula != o_formula:
                    stats.add(Finding(
                        CRITICAL, "FORMULA_MISMATCH", sheet_name, coord,
                        f"Formula text differs.\n"
                        f"    Master formula: {m_formula!r}\n"
                        f"    Output formula: {o_formula!r}\n"
                        f"This indicates a formula-driven cell was altered "
                        f"outside the two/three legitimate write-points -- "
                        f"this must NEVER happen in a production output.",
                        master_value=m_formula, output_value=o_formula,
                    ))
            else:
                if m_cell.value != o_cell.value:
                    stats.add(Finding(
                        CRITICAL, "STATIC_CONTENT_MISMATCH", sheet_name, coord,
                        f"Static (non-formula) content differs.\n"
                        f"    Master value: {m_cell.value!r}\n"
                        f"    Output value: {o_cell.value!r}\n"
                        f"This cell is outside every recognized input "
                        f"region -- it should be an immutable label, "
                        f"heading, or constant, and it changed anyway.",
                        master_value=m_cell.value, output_value=o_cell.value,
                    ))


# ---------------------------------------------------------------------------
# Report printing
# ---------------------------------------------------------------------------
def print_report(stats: RunStats, master_path: str, output_path: str) -> None:
    print("\n" + "=" * 100)
    print("FORMULA INTEGRITY GUARD -- REPORT")
    print("=" * 100)
    print(f"Master template : {master_path}")
    print(f"Output file     : {output_path}")
    print(f"Sheets compared  : {stats.sheets_compared}")
    print(f"Cells scanned    : {stats.cells_scanned}")
    print(f"Cells exempt     : {stats.cells_exempt}")
    print(f"Total findings   : {len(stats.findings)} "
          f"({len(stats.criticals())} CRITICAL, {len(stats.warnings())} WARNING)")
    print("-" * 100)

    if not stats.findings:
        print("No discrepancies of any kind detected. Output file is structurally "
              "and formulaically identical to the master template outside the "
              "recognized input regions.")
        print("=" * 100)
        return

    by_category: dict = {}
    for f in stats.findings:
        by_category.setdefault(f.category, []).append(f)

    for category, findings in sorted(by_category.items()):
        severity = findings[0].severity
        print(f"\n[{severity}] {category} -- {len(findings)} finding(s)")
        print("-" * 100)
        for f in findings:
            print(f"  Sheet: {f.sheet:30} Cell: {f.cell}")
            print(f"    {f.message}")
            print()

    print("=" * 100)
    if stats.criticals():
        print(f"RESULT: FAIL -- {len(stats.criticals())} CRITICAL finding(s). "
              f"DO NOT ship this output file.")
    else:
        print(f"RESULT: PASS WITH WARNINGS -- {len(stats.warnings())} warning(s), "
              f"no critical findings.")
    print("=" * 100)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------
def run_guard(master_path: str, output_path: str,
              check_formatting: bool = False,
              check_defined_names_flag: bool = True,
              treat_warnings_as_failure: bool = False,
              expected_master_sha256: Optional[str] = None) -> int:

    verify_master_checksum(master_path, expected_master_sha256)

    wb_master = load_workbook_or_die(master_path, "master template")
    wb_output = load_workbook_or_die(output_path, "output")

    stats = RunStats()

    exempt = build_exempt_cells(wb_master)

    common_sheets = check_sheet_structure(wb_master, wb_output, stats)
    stats.sheets_compared = len(common_sheets)

    if check_defined_names_flag:
        check_defined_names(wb_master, wb_output, stats)

    for sheet_name in common_sheets:
        ws_master = wb_master[sheet_name]
        ws_output = wb_output[sheet_name]

        sheet_exempt = exempt.get(sheet_name, set())

        compare_sheet_cells(ws_master, ws_output, sheet_name, sheet_exempt, stats)
        check_merged_ranges(ws_master, ws_output, sheet_name, stats)

        if check_formatting:
            max_row = max(ws_master.max_row, ws_output.max_row)
            max_col = max(ws_master.max_column, ws_output.max_column)
            check_number_formats(ws_master, ws_output, sheet_name, sheet_exempt,
                                  max_row, max_col, stats)

    wb_master.close()
    wb_output.close()

    print_report(stats, master_path, output_path)

    if stats.criticals():
        return 1
    if stats.warnings() and treat_warnings_as_failure:
        return 1
    return 0


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=(
            "Pre-deployment integrity guard: verifies a generated financial "
            "statement output file has not deviated from the master "
            "template anywhere outside its legitimate input regions."
        )
    )
    parser.add_argument("master_path", help="Path to the pristine master template.")
    parser.add_argument("output_path", help="Path to the generated output file to verify.")
    parser.add_argument(
        "--check-formatting", action="store_true",
        help="Also compare number_format strings on non-input cells (warning-level only).",
    )
    parser.add_argument(
        "--no-check-defined-names", action="store_true",
        help="Skip the workbook-level defined-names comparison (enabled by default).",
    )
    parser.add_argument(
        "--strict", action="store_true",
        help="Treat warning-level findings as build failures too (exit code 1).",
    )
    parser.add_argument(
        "--expected-master-sha256", default=None,
        help=(
            "Optional known-good SHA-256 checksum of the master template. "
            "If provided, the tool refuses to proceed if the file at "
            "master_path does not match it -- guards against accidentally "
            "running this check against a tampered or wrong master file."
        ),
    )

    args = parser.parse_args()

    exit_code = run_guard(
        master_path=args.master_path,
        output_path=args.output_path,
        check_formatting=args.check_formatting,
        check_defined_names_flag=not args.no_check_defined_names,
        treat_warnings_as_failure=args.strict,
        expected_master_sha256=args.expected_master_sha256,
    )

    sys.exit(exit_code)
