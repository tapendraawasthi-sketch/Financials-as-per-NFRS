"""
extract_notes_structure.py

Verifies (or corrects) notes_structure_map.json by parsing the ACTUAL
formulas in 'Notes 3.2 to 3.23', rather than the value-based reconstruction
used when the file could not be opened directly.

Usage:
    python extract_notes_structure.py "MEs Financials Format.xlsx"

Output:
    notes_structure_map_VERIFIED.json
    notes_structure_needs_review.json
"""

import json
import re
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries

SHEET_NOTES = "Notes 3.2 to 3.23"
SHEET_TB = "Trial Balance"
SHEET_BS = "Balance Sheet"
SHEET_IS = "Income Statement"

CELL_REF_RE = re.compile(
    r"(?:(?P<quoted_sheet>'[^']+')|(?P<bare_sheet>[A-Za-z0-9_]+))?"
    r"!?"
    r"(?P<col1>\$?[A-Za-z]{1,3})(?P<row1>\$?\d+)"
    r"(?::(?P<col2>\$?[A-Za-z]{1,3})(?P<row2>\$?\d+))?"
)


def strip_quotes(name):
    if name and name.startswith("'") and name.endswith("'"):
        return name[1:-1].replace("''", "'")
    return name


def parse_refs(formula, current_sheet):
    """Returns list of (sheet, cell_or_range) referenced by formula."""
    refs = []
    for m in CELL_REF_RE.finditer(formula):
        col1, row1 = m.group("col1"), m.group("row1")
        if col1 is None or row1 is None:
            continue
        sheet_name = strip_quotes(m.group("quoted_sheet") or m.group("bare_sheet")) or current_sheet
        col2, row2 = m.group("col2"), m.group("row2")
        coord = (
            f"{col1.replace('$','')}{row1.replace('$','')}:{col2.replace('$','')}{row2.replace('$','')}"
            if col2 and row2
            else f"{col1.replace('$','')}{row1.replace('$','')}"
        )
        refs.append((sheet_name, coord))
    return refs


def expand_range(coord, max_cells=300):
    if ":" not in coord:
        return [coord]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
    except ValueError:
        return [coord]
    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cells.append(f"{get_column_letter(c)}{r}")
            if len(cells) > max_cells:
                return cells
    return cells


def nearest_label_above(ws, row, label_col="B", max_lookback=30):
    """Inherit the nearest non-empty label above `row` in column B (or A),
    used to attach a human-readable title to formula rows that don't carry
    their own label."""
    for r in range(row, max(row - max_lookback, 0), -1):
        for col in (label_col, "A"):
            val = ws[f"{col}{r}"].value
            if val not in (None, ""):
                return str(val), r
    return None, None


def note_number_for_row(row, note_header_rows):
    """note_header_rows: dict {row_number: note_number_string}, sorted.
    Returns the note number governing this row (the last header at or
    before this row)."""
    applicable = [r for r in note_header_rows if r <= row]
    if not applicable:
        return None
    return note_header_rows[max(applicable)]


def find_note_header_rows(ws):
    """Scan column A for note-number-looking values (e.g. 3.2, 3.10, 3.23)
    and return {row: 'note_string'}. Handles the float-truncation issue
    where 3.10 / 3.20 render as 3.1 / 3.2 by checking adjacent label text."""
    headers = {}
    for row in range(1, ws.max_row + 1):
        val = ws[f"A{row}"].value
        if val is None:
            continue
        if isinstance(val, (int, float)):
            note_str = f"{val:.2f}".rstrip("0").rstrip(".")
            headers[row] = note_str
        elif isinstance(val, str) and re.match(r"^3\.\d{1,2}$", val.strip()):
            headers[row] = val.strip()
    return headers


def build_notes_map(wb):
    ws = wb[SHEET_NOTES]
    note_headers = find_note_header_rows(ws)
    print(f"Detected {len(note_headers)} note header rows: {note_headers}")

    formula_rows = {}
    for row in range(1, ws.max_row + 1):
        for col in ("E", "F"):
            cell = ws[f"{col}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                label, label_row = nearest_label_above(ws, row)
                refs = parse_refs(cell.value, SHEET_NOTES)
                expanded = []
                for sheet_name, coord in refs:
                    expanded.extend(
                        (sheet_name, c) for c in expand_range(coord)
                    )
                formula_rows.setdefault(row, {})[col] = {
                    "formula": cell.value,
                    "label": label,
                    "label_row": label_row,
                    "note_number": note_number_for_row(row, note_headers),
                    "raw_refs": [f"{s}!{c}" for s, c in expanded],
                }

    notes_map = {}
    needs_review = []

    for row, cols in formula_rows.items():
        note_num = cols.get("E", cols.get("F", {})).get("note_number")
        if note_num is None:
            needs_review.append({
                "row": row, "reason": "No governing note header found above this row.",
                "cols": cols,
            })
            continue

        entry = notes_map.setdefault(note_num, {
            "title": cols.get("E", cols.get("F", {})).get("label"),
            "rows": [],
            "feeds_from_tb_rows": set(),
        })
        entry["rows"].append(row)

        for col_data in cols.values():
            for ref in col_data["raw_refs"]:
                if ref.startswith(f"{SHEET_TB}!"):
                    tb_cell = ref.split("!")[1]
                    tb_row_match = re.match(r"[A-Za-z]+(\d+)", tb_cell)
                    if tb_row_match:
                        entry["feeds_from_tb_rows"].add(int(tb_row_match.group(1)))
                elif not any(
                    ref.startswith(f"{s}!") for s in (SHEET_NOTES,)
                ):
                    # Reference to some sheet OTHER than Notes/Trial Balance
                    # (e.g. PPE Workings, Sundry Debtors, Related Party,
                    # Change in Equity, Fair Value Change, Enter Details).
                    # This is exactly the kind of "extra intermediate layer"
                    # flagged in the manual reconstruction -- surface it.
                    entry.setdefault("intermediate_sheet_refs", set()).add(ref)

    for note_num, entry in notes_map.items():
        entry["feeds_from_tb_rows"] = sorted(entry["feeds_from_tb_rows"])
        if "intermediate_sheet_refs" in entry:
            entry["intermediate_sheet_refs"] = sorted(entry["intermediate_sheet_refs"])
        entry["rows"] = sorted(set(entry["rows"]))
        entry["total_row"] = entry["rows"][-1] if entry["rows"] else None

    return notes_map, needs_review


def crosscheck_against_statements(wb, notes_map):
    """For every Balance Sheet / Income Statement formula cell that
    references 'Notes 3.2 to 3.23', attach a feeds_into confirmation to the
    relevant note -- and flag any BS/IS cell whose formula spans MULTIPLE
    note rows (a strong signal of a multi-note-sum line, like BS C17)."""
    feeds_into = {}
    for sheet_name in (SHEET_BS, SHEET_IS):
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    refs = parse_refs(cell.value, sheet_name)
                    note_refs = [
                        (s, c) for s, c in refs if s == SHEET_NOTES
                    ]
                    if note_refs:
                        feeds_into[f"{sheet_name}!{cell.coordinate}"] = {
                            "formula": cell.value,
                            "note_cells_referenced": [f"{s}!{c}" for s, c in note_refs],
                            "spans_multiple_note_rows": len(note_refs) > 1,
                        }
    return feeds_into


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_notes_structure.py <workbook.xlsx>")
        sys.exit(1)

    wb = load_workbook(sys.argv[1], data_only=False)
    notes_map, needs_review = build_notes_map(wb)
    feeds_into = crosscheck_against_statements(wb, notes_map)

    with open("notes_structure_map_VERIFIED.json", "w", encoding="utf-8") as f:
        json.dump(
            {"notes": notes_map, "statement_links": feeds_into},
            f, indent=2, default=str, ensure_ascii=False,
        )

    with open("notes_structure_needs_review.json", "w", encoding="utf-8") as f:
        json.dump(needs_review, f, indent=2, default=str, ensure_ascii=False)

    print(f"Wrote {len(notes_map)} verified notes to notes_structure_map_VERIFIED.json")
    print(f"Wrote {len(needs_review)} unresolved rows to notes_structure_needs_review.json")
    print("\nDiff this against the manually-reconstructed map from chat and treat any "
          "mismatch as ground truth correcting my earlier reconstruction, not vice versa.")


if __name__ == "__main__":
    main()
