"""
test_full_chain.py

Regression test proving that a value written at Layer 1 (Trial Balance)
propagates correctly through Layer 2 ("Notes 3.2 to 3.23") to Layer 3
(Balance Sheet / Income Statement) -- using the REAL template's formulas
via a real recalculation engine (LibreOffice headless), not a hand-rolled
reimplementation of the formula logic.

Three chain lengths are exercised:
  A. Paid-up Capital        (TB row 7)   -> Note 3.9  -> Balance Sheet!C24   (see caveat below)
  B. Provision for CSR      (TB row 27)  -> Note 3.15 -> Balance Sheet!C38
  C. Interest expense       (TB row 110) -> Income Statement!C14 DIRECTLY (confirmed layer-skip)

CAVEAT on case A: the "Change in Equity -> Balance Sheet" framing this test
was originally requested against does not match this project's own
notes_structure_map.json reconstruction, which routes Balance Sheet!C24
through Note 3.9 instead. This test checks the Balance Sheet outcome as
the authoritative assertion and treats Change in Equity as a secondary,
informational check -- run it and let the numbers settle the question.

This file PREFERS importing the real production functions from
excel_writer_service.py (safe_write_cell, recalculate_workbook). If that
module isn't importable, it falls back to local reimplementations of
identical logic, clearly flagged as less authoritative.
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Prefer the real production writer/recalc functions.
# ---------------------------------------------------------------------------
try:
    from excel_writer_service import (
        safe_write_cell as _prod_safe_write_cell,
        recalculate_workbook as _prod_recalculate_workbook,
        SHEET_TRIAL_BALANCE,
    )
    _USING_PRODUCTION_WRITER = True
except ImportError:
    _USING_PRODUCTION_WRITER = False
    SHEET_TRIAL_BALANCE = "Trial Balance"

    def _fallback_safe_write_cell(wb, sheet: str, coord: str, value, whitelist: set[str]) -> None:
        key = f"{sheet}!{coord}"
        if key not in whitelist:
            raise PermissionError(
                f"Illegal write blocked: '{key}' is not in the input whitelist "
                f"(local fallback safe_write_cell -- excel_writer_service.py "
                f"was not importable)."
            )
        wb[sheet][coord] = value

    def _fallback_recalculate_workbook(file_path: str, timeout_seconds: int = 120) -> bool:
        file_path = str(file_path)
        with tempfile.TemporaryDirectory() as tmp_dir:
            try:
                subprocess.run(
                    [
                        "soffice", "--headless", "--norestore",
                        "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                        "--outdir", tmp_dir, file_path,
                    ],
                    capture_output=True, text=True, timeout=timeout_seconds, check=True,
                )
                converted = Path(tmp_dir) / (Path(file_path).stem + ".xlsx")
                if not converted.exists():
                    return False
                shutil.move(str(converted), file_path)
                return True
            except (FileNotFoundError, subprocess.TimeoutExpired, subprocess.CalledProcessError) as exc:
                print(f"[WARNING] LibreOffice recalculation failed/unavailable: {exc}")
                return False

    _prod_safe_write_cell = _fallback_safe_write_cell
    _prod_recalculate_workbook = _fallback_recalculate_workbook


# ---------------------------------------------------------------------------
# Structural whitelist: known legitimate Trial Balance input columns.
# Rows 6-149, current-year D/E/F/G, previous-year During-only N/O (per the
# original spec: previous year gets no Adjustment columns).
# ---------------------------------------------------------------------------
TB_ROW_MIN, TB_ROW_MAX = 6, 149
TB_INPUT_COLS_CY = ["D", "E", "F", "G"]
TB_INPUT_COLS_PY = ["N", "O"]


def _load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_structural_trial_balance_whitelist() -> set[str]:
    cells = set()
    for row in range(TB_ROW_MIN, TB_ROW_MAX + 1):
        for col in TB_INPUT_COLS_CY + TB_INPUT_COLS_PY:
            cells.add(f"{SHEET_TRIAL_BALANCE}!{col}{row}")
    return cells


def load_input_whitelist(graph_path: str) -> set[str]:
    """
    Builds the writable-cell whitelist. Uses BOTH:
      (1) the structural rule (TB D/E/F/G/N/O within rows 6-149), and
      (2) any explicit is_green_input_candidate flags in the graph, if present.
    Then REMOVES any candidate cell that the dependency graph shows already
    contains a formula -- this is the critical safety net, since several TB
    During/Adjustment cells are auto-linked from Sundry Debtors/Creditors,
    Bank Acc., Related Party, PPE Workings, Fair Value Change, or Tax
    Calculation sheets (per the Instructions sheet), and are therefore NOT
    legitimate raw-input cells despite sitting inside the expected column
    range.
    """
    graph = _load_json(graph_path)
    entries = graph.get("dependency_graph", graph)

    structural = build_structural_trial_balance_whitelist()

    flagged = set()
    conflicts = []
    for key, meta in entries.items():
        is_green = bool(meta.get("is_green_input_candidate", False))
        has_formula = bool(meta.get("raw_formula"))
        if is_green and has_formula:
            conflicts.append(key)
            continue
        if is_green and not has_formula:
            flagged.add(key)

    whitelist = structural | flagged

    removed = []
    for key in list(whitelist):
        meta = entries.get(key)
        if meta and meta.get("raw_formula"):
            whitelist.discard(key)
            removed.append(key)

    if conflicts:
        print(f"[WARNING] {len(conflicts)} green-filled cell(s) ALSO contain a "
              f"formula and were excluded from the whitelist: {sorted(conflicts)}")
    if removed:
        print(f"[WARNING] {len(removed)} cell(s) inside the expected Trial Balance "
              f"input range are actually formula-driven (auto-linked from a "
              f"sub-schedule sheet) and have been REMOVED from the whitelist:")
        for r in sorted(removed):
            print(f"    - {r}  (raw_formula: {entries[r].get('raw_formula')!r})")

    print(f"[INFO] Final whitelist size: {len(whitelist)} cells.")
    return whitelist


def preflight_check_not_formula(graph: dict, sheet: str, coord: str) -> None:
    """Aborts loudly if a chosen test-write cell is actually formula-driven."""
    key = f"{sheet}!{coord}"
    entries = graph.get("dependency_graph", graph)
    meta = entries.get(key)
    if meta and meta.get("raw_formula"):
        raise RuntimeError(
            f"PRE-FLIGHT CHECK FAILED for {key}: the dependency graph shows "
            f"this cell already contains a formula ({meta['raw_formula']!r}). "
            f"This test case's assumption that {key} is a genuine raw-input "
            f"cell is WRONG -- pick a different account before proceeding."
        )


# ---------------------------------------------------------------------------
# Test case definitions
# ---------------------------------------------------------------------------
@dataclass
class ChainTestCase:
    case_id: str
    description: str
    tb_row: int
    tb_col: str
    delta: float
    target_sheet: str
    target_cell: str
    chain_note: str
    secondary_check: Optional[dict] = None


TEST_CASES = [
    ChainTestCase(
        case_id="A_simple_passthrough",
        description=(
            "Paid-up Capital: TB row 7, During Cr (+2,000,000) -> "
            "Note 3.9 (Share Capital) -> Balance Sheet!C24"
        ),
        tb_row=7, tb_col="E", delta=2_000_000.0,
        target_sheet="Balance Sheet", target_cell="C24",
        chain_note=(
            "The original framing described this as TB -> Change in Equity -> "
            "Balance Sheet!C24 (2-hop). Per this project's own "
            "notes_structure_map.json reconstruction, Balance Sheet!C24 is "
            "actually fed via Note 3.9 inside 'Notes 3.2 to 3.23' (a 3-hop "
            "chain), and 'Change in Equity' is an INDEPENDENT downstream "
            "consumer of the same Trial Balance write, not an intermediate "
            "hop feeding the Balance Sheet. The primary assertion here checks "
            "the Balance Sheet outcome directly; Change in Equity is checked "
            "as a secondary/informational assertion so this run empirically "
            "settles which framing is correct."
        ),
        secondary_check={
            "sheet": "Change in Equity",
            "cell": "B18",
            "label": (
                "Issue of Share Capital, Share Capital column, current-year "
                "section -- COORDINATE UNVERIFIED, confirm against the real "
                "file before trusting this specific cell reference."
            ),
        },
    ),
    ChainTestCase(
        case_id="B_notes_routed",
        description=(
            "Provision for CSR: TB row 27, During Cr (+15,000) -> "
            "Note 3.15 (Provisions) -> Balance Sheet!C38"
        ),
        tb_row=27, tb_col="E", delta=15_000.0,
        target_sheet="Balance Sheet", target_cell="C38",
        chain_note=(
            "Chosen from this project's notes_structure_map.json "
            "reconstruction (Note 3.15, total_row 146, feeds_into "
            "Balance Sheet!C38). This is the 'textbook' 3-hop case: a single "
            "TB row feeding a simple SUM note total feeding a single "
            "statement line, with no cross-section merges (unlike, e.g., the "
            "Wages/Note 3.20 case, which was deliberately avoided here "
            "because it would make hand-verification ambiguous)."
        ),
    ),
    ChainTestCase(
        case_id="C_direct_skip",
        description=(
            "Interest expense: TB row 110, During Dr (+40,000) -> "
            "Income Statement!C14 DIRECTLY (bypassing Notes 3.2-3.23)"
        ),
        tb_row=110, tb_col="D", delta=40_000.0,
        target_sheet="Income Statement", target_cell="C14",
        chain_note=(
            "This is the confirmed layer-skip case from the original chain "
            "analysis: Income Statement!C14 = 'Trial Balance'!I110 directly, "
            "with NO row in 'Notes 3.2 to 3.23' involved. This is the "
            "highest-risk pattern for a writer-service bug (accidentally "
            "routing a value through Notes when the real template doesn't), "
            "so it gets its own dedicated test."
        ),
    ),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_cell_value(path, sheet: str, coord: str, data_only: bool):
    wb = load_workbook(path, data_only=data_only)
    try:
        return wb[sheet][coord].value
    finally:
        wb.close()


def to_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    raise ValueError(f"Expected a numeric cell value, got {val!r} ({type(val).__name__})")


def write_cell(wb, sheet: str, coord: str, value, whitelist: set[str]) -> None:
    _prod_safe_write_cell(wb, sheet, coord, value, whitelist)


def recalc(path) -> bool:
    return _prod_recalculate_workbook(str(path))


def run_corruption_check(path, graph: dict, whitelist: set[str]) -> list[dict]:
    """
    Opens the post-write, post-recalc workbook with data_only=False and
    checks EVERY cell in the dependency graph that originally had a formula
    and is NOT in the whitelist -- asserting it STILL contains a formula
    string. Any violation here means a write leaked outside the two/three
    legitimate write-points and clobbered a formula cell.
    """
    wb = load_workbook(path, data_only=False)
    entries = graph.get("dependency_graph", graph)
    violations = []
    try:
        for key, meta in entries.items():
            if not meta.get("raw_formula"):
                continue
            if key in whitelist:
                continue
            sheet, coord = key.split("!", 1)
            if sheet not in wb.sheetnames:
                continue
            current_value = wb[sheet][coord].value
            still_formula = isinstance(current_value, str) and current_value.startswith("=")
            if not still_formula:
                violations.append({
                    "cell": key,
                    "expected_formula": meta["raw_formula"],
                    "actual_value": current_value,
                })
    finally:
        wb.close()
    return violations


def walk_chain(graph: dict, key: str, depth: int = 0, max_depth: int = 12,
                visited: Optional[set] = None, lines: Optional[list] = None) -> list[str]:
    """Recursively renders the reference chain for `key` using the flat
    schema {"raw_formula": str, "direct_refs": [str], "sheet": str,
    "row_label": str} -- cycle-safe, depth-capped."""
    if visited is None:
        visited = set()
    if lines is None:
        lines = []
    indent = "  " * depth

    if key in visited:
        lines.append(f"{indent}-> {key}  [CYCLE DETECTED -- stopping this branch]")
        return lines
    visited.add(key)

    entries = graph.get("dependency_graph", graph)
    meta = entries.get(key)
    if meta is None:
        lines.append(f"{indent}-> {key}  [NOT FOUND in dependency graph]")
        return lines

    label = meta.get("row_label") or ""
    formula = meta.get("raw_formula")
    lines.append(f"{indent}-> {key}  ({label})  formula={formula!r}")

    if depth >= max_depth:
        lines.append(f"{indent}   [max recursion depth reached -- truncating]")
        return lines

    for ref in meta.get("direct_refs", []) or []:
        walk_chain(graph, ref, depth + 1, max_depth, visited, lines)

    return lines


def print_failure_chain(graph: dict, notes_map: Optional[dict], case: ChainTestCase) -> None:
    print(f"\n--- Full reference chain for failing/blocked case: {case.case_id} ---")
    target_key = f"{case.target_sheet}!{case.target_cell}"
    for line in walk_chain(graph, target_key):
        print(line)

    if notes_map:
        notes = notes_map.get("notes", notes_map)
        print("\n  Related notes_structure_map.json entries (matched by cell/sheet mention):")
        found_any = False
        for note_num, entry in notes.items():
            entry_str = json.dumps(entry, default=str)
            if case.target_cell in entry_str or case.target_sheet in entry_str:
                found_any = True
                print(f"    Note {note_num}: {json.dumps(entry, indent=2, default=str)}")
        if not found_any:
            print("    (no matching note entries found for this target cell)")

    print(f"\n  Design note for this case:\n  {case.chain_note}")


def print_results_table(results: list[dict]) -> None:
    print("\n" + "=" * 108)
    print(f"{'Test Case':32} | {'Expected':>16} | {'Actual':>16} | {'Result':>14}")
    print("-" * 108)
    for r in results:
        case = r["case"]
        status = r["status"]
        if status in ("BLOCKED_BY_WHITELIST", "INCONCLUSIVE"):
            print(f"{case.case_id:32} | {'N/A':>16} | {'N/A':>16} | {status:>14}")
            if r.get("error"):
                print(f"    reason: {r['error']}")
            continue

        print(f"{case.case_id:32} | {r['expected']:>16,.2f} | {r['actual']:>16,.2f} | {status:>14}")
        if r.get("corruption_violations"):
            print(f"    !! {len(r['corruption_violations'])} corruption violation(s) detected outside the whitelist:")
            for v in r["corruption_violations"][:10]:
                print(f"       {v['cell']}: expected formula {v['expected_formula']!r}, "
                      f"found value {v['actual_value']!r}")
            if len(r["corruption_violations"]) > 10:
                print(f"       ... and {len(r['corruption_violations']) - 10} more.")
        if r.get("secondary_check_result"):
            sc = r["secondary_check_result"]
            flag = "OK" if sc["match"] else "MISMATCH"
            print(f"    (secondary/informational -- {sc['sheet_cell']}: expected "
                  f"{sc['expected']:,.2f}, actual {sc['actual']:,.2f} -> {flag})")
            print(f"       {sc['note']}")
    print("=" * 108)


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------
def run_all(template_path: str, graph_path: str, notes_map_path: Optional[str],
            tolerance: float = 1.0) -> dict:
    print(f"[INFO] Production writer service "
          f"{'FOUND -- exercising the real safe_write_cell/recalculate_workbook.' if _USING_PRODUCTION_WRITER else 'NOT FOUND -- using local fallback reimplementations (less authoritative).'}")

    graph = _load_json(graph_path)
    notes_map = _load_json(notes_map_path) if notes_map_path else None
    whitelist = load_input_whitelist(graph_path)

    for case in TEST_CASES:
        preflight_check_not_formula(graph, SHEET_TRIAL_BALANCE, f"{case.tb_col}{case.tb_row}")
    print("[INFO] Pre-flight check passed: all three chosen TB write-cells are "
          "confirmed non-formula in the dependency graph.")

    with tempfile.TemporaryDirectory() as tmp_dir_str:
        tmp_dir = Path(tmp_dir_str)
        pristine_path = tmp_dir / "pristine_recalculated.xlsx"
        shutil.copy2(template_path, pristine_path)

        print("\n[INFO] Recalculating a pristine template copy to establish baselines...")
        if not recalc(pristine_path):
            print("\n" + "=" * 78)
            print("ABORTING: LibreOffice headless recalculation is unavailable or failed.")
            print("Without a real recalculation engine, 'actual' values read back would")
            print("be STALE CACHED VALUES from whenever the file was last saved -- ")
            print("comparing against those would be meaningless and could produce a")
            print("false PASS. Marking the entire suite INCONCLUSIVE.")
            print("=" * 78)
            return {"status": "INCONCLUSIVE", "results": []}

        baselines = {}
        for case in TEST_CASES:
            baselines[case.case_id] = to_number(
                get_cell_value(pristine_path, case.target_sheet, case.target_cell, data_only=True)
            )
            if case.secondary_check:
                baselines[case.case_id + "__secondary"] = to_number(
                    get_cell_value(pristine_path, case.secondary_check["sheet"],
                                    case.secondary_check["cell"], data_only=True)
                )

        results = []
        for case in TEST_CASES:
            case_path = tmp_dir / f"case_{case.case_id}.xlsx"
            shutil.copy2(pristine_path, case_path)

            result = {"case": case}

            try:
                wb = load_workbook(case_path, data_only=False)
                coord = f"{case.tb_col}{case.tb_row}"
                old_val = to_number(wb[SHEET_TRIAL_BALANCE][coord].value)
                new_val = old_val + case.delta
                write_cell(wb, SHEET_TRIAL_BALANCE, coord, new_val, whitelist)
                wb.save(case_path)
                wb.close()
            except PermissionError as exc:
                result.update({"status": "BLOCKED_BY_WHITELIST", "error": str(exc)})
                results.append(result)
                continue

            if not recalc(case_path):
                result.update({"status": "INCONCLUSIVE",
                                "error": "Recalculation failed for this case."})
                results.append(result)
                continue

            actual = to_number(get_cell_value(case_path, case.target_sheet, case.target_cell, data_only=True))
            expected = baselines[case.case_id] + case.delta
            value_match = abs(actual - expected) <= tolerance

            violations = run_corruption_check(case_path, graph, whitelist)

            result.update({
                "status": "PASS" if (value_match and not violations) else "FAIL",
                "old_val": old_val, "new_val": new_val, "delta": case.delta,
                "baseline": baselines[case.case_id], "expected": expected, "actual": actual,
                "value_match": value_match, "corruption_violations": violations,
            })

            if case.secondary_check:
                sec_actual = to_number(get_cell_value(
                    case_path, case.secondary_check["sheet"], case.secondary_check["cell"], data_only=True))
                sec_expected = baselines[case.case_id + "__secondary"] + case.delta
                result["secondary_check_result"] = {
                    "sheet_cell": f"{case.secondary_check['sheet']}!{case.secondary_check['cell']}",
                    "expected": sec_expected, "actual": sec_actual,
                    "match": abs(sec_actual - sec_expected) <= tolerance,
                    "note": "Informational only -- does not affect this case's overall PASS/FAIL.",
                }

            results.append(result)

        print_results_table(results)

        for result in results:
            if result["status"] in ("FAIL", "BLOCKED_BY_WHITELIST", "INCONCLUSIVE"):
                print_failure_chain(graph, notes_map, result["case"])

        overall_status = "PASS" if all(r["status"] == "PASS" for r in results) else "FAIL"
        return {"status": overall_status, "results": results}


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Regression test: Trial Balance write -> Notes 3.2-3.23 -> "
                     "Balance Sheet/Income Statement propagation, using a real "
                     "recalculation engine."
    )
    parser.add_argument("--template", default="MEs Financials Format.xlsx")
    parser.add_argument("--graph", default="formula_dependency_graph.json")
    parser.add_argument("--notes-map", default="notes_structure_map.json")
    parser.add_argument("--tolerance", type=float, default=1.0,
                         help="Absolute NPR tolerance for value comparisons (default 1.0).")
    args = parser.parse_args()

    if not Path(args.template).exists():
        print(f"FATAL: template not found at {args.template}")
        sys.exit(2)
    if not Path(args.graph).exists():
        print(f"FATAL: dependency graph not found at {args.graph}")
        sys.exit(2)

    notes_map_path = args.notes_map if Path(args.notes_map).exists() else None
    if notes_map_path is None:
        print(f"[WARNING] notes_structure_map.json not found at {args.notes_map} -- "
              f"failure chain output will skip the notes-map cross-reference section.")

    outcome = run_all(args.template, args.graph, notes_map_path, tolerance=args.tolerance)

    print(f"\nOVERALL SUITE STATUS: {outcome['status']}")
    sys.exit(0 if outcome["status"] == "PASS" else 1)
