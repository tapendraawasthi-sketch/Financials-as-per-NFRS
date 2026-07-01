"""
extract_full_formula_chain.py

Builds a full formula dependency graph for the MEs Financials workbook.

Usage:
    python extract_full_formula_chain.py "MEs Financials Format.xlsx"

Output:
    formula_dependency_graph.json
"""

import json
import re
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries

# ---------------------------------------------------------------------------
# Regex to pull cell / range references out of a formula string.
# Handles:
#   'Sheet Name'!A1          (quoted sheet, spaces allowed)
#   SheetName!A1              (unquoted sheet, no spaces)
#   A1                        (same-sheet reference)
#   A1:B5 / Sheet!A1:B5       (ranges)
#   $A$1 style absolute refs
# Does NOT attempt to resolve defined names, table references, or
# structured references -- those are reported as "unresolved_tokens".
# ---------------------------------------------------------------------------
CELL_REF_RE = re.compile(
    r"(?:(?P<quoted_sheet>'[^']+')|(?P<bare_sheet>[A-Za-z0-9_]+))?"
    r"!?"
    r"(?P<col1>\$?[A-Za-z]{1,3})(?P<row1>\$?\d+)"
    r"(?::(?P<col2>\$?[A-Za-z]{1,3})(?P<row2>\$?\d+))?"
)

GREEN_FILL_HINTS = {
    "FF92D050", "FFC6E0B4", "FF00B050", "FFA9D08E", "FFE2EFDA", "FF375623"
}


def is_greenish(rgb):
    """Heuristic: treat as a 'green input' cell if fill exists and the
    green channel clearly dominates red/blue, OR it matches a known
    template green hex. rgb is an 8-char ARGB hex string or None."""
    if not rgb or not isinstance(rgb, str) or len(rgb) != 8:
        return False
    if rgb.upper() in GREEN_FILL_HINTS:
        return True
    try:
        r = int(rgb[2:4], 16)
        g = int(rgb[4:6], 16)
        b = int(rgb[6:8], 16)
    except ValueError:
        return False
    return g > r + 15 and g > b + 15 and g > 100


def strip_sheet_quotes(name):
    if name and name.startswith("'") and name.endswith("'"):
        return name[1:-1].replace("''", "'")
    return name


def normalize_ref(sheet, coord):
    return f"{sheet}!{coord}"


def find_refs_in_formula(formula, current_sheet):
    """Extract a list of (sheet, cell_or_range) tuples referenced by a
    formula string. Cross-sheet refs use the named sheet; bare refs use
    current_sheet. Ranges are expanded to a bounding description (not
    exploded into every cell -- expansion happens lazily in resolve_to_source
    only when a range is small enough to be worth walking)."""
    refs = []
    for m in CELL_REF_RE.finditer(formula):
        col1, row1 = m.group("col1"), m.group("row1")
        if col1 is None or row1 is None:
            continue
        sheet_name = m.group("quoted_sheet") or m.group("bare_sheet")
        sheet_name = strip_sheet_quotes(sheet_name) if sheet_name else current_sheet
        col2, row2 = m.group("col2"), m.group("row2")
        if col2 and row2:
            coord = f"{col1.replace('$','')}{row1.replace('$','')}:{col2.replace('$','')}{row2.replace('$','')}"
        else:
            coord = f"{col1.replace('$','')}{row1.replace('$','')}"
        refs.append((sheet_name, coord))
    return refs


def expand_range_if_small(sheet_name, coord, max_cells=200):
    """If coord is a range (A1:B5), expand to individual cell coords,
    capped at max_cells to avoid pathological blow-ups on full-column refs."""
    if ":" not in coord:
        return [coord]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
    except ValueError:
        return [coord]  # can't parse, return as-is
    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cells.append(f"{get_column_letter(c)}{r}")
            if len(cells) > max_cells:
                return cells  # bail out, too big to be a normal single-cell link
    return cells


def get_row_label(ws, row_idx, label_cols=(1, 2)):
    """Best-effort label lookup: look in column A or B of the same row."""
    for col in label_cols:
        try:
            val = ws.cell(row=row_idx, column=col).value
        except Exception:
            continue
        if val not in (None, ""):
            return str(val)
    return None


def resolve_to_source(workbook, sheet, cell, visited=None):
    """Recursively resolve `sheet!cell` down to its deepest non-formula
    source cell(s). Returns a list of 'Sheet!Cell' strings representing
    the leaves of the dependency chain. Cycle-safe via `visited`."""
    if visited is None:
        visited = set()

    key = (sheet, cell)
    if key in visited:
        print(f"[WARNING] Circular reference detected at {sheet}!{cell} -- "
              f"stopping recursion for this branch.")
        return [f"{sheet}!{cell} (CYCLE)"]
    visited.add(key)

    if sheet not in workbook.sheetnames:
        return [f"{sheet}!{cell} (SHEET NOT FOUND)"]

    ws = workbook[sheet]
    try:
        cell_obj = ws[cell.split(":")[0]]  # if a range slipped through, use first cell
    except Exception:
        return [f"{sheet}!{cell} (INVALID CELL)"]

    value = cell_obj.value

    if not (isinstance(value, str) and value.startswith("=")):
        # Not a formula -> this IS a source cell.
        return [f"{sheet}!{cell}"]

    formula = value
    refs = find_refs_in_formula(formula, current_sheet=sheet)

    if not refs:
        # Formula exists but no cell refs found (e.g. pure constant formula
        # like "=1+1" or a function we didn't parse) -> treat as source.
        return [f"{sheet}!{cell} (FORMULA WITH NO RESOLVABLE REFS: {formula})"]

    leaves = []
    for ref_sheet, ref_coord in refs:
        for single_cell in expand_range_if_small(ref_sheet, ref_coord):
            leaves.extend(resolve_to_source(workbook, ref_sheet, single_cell, visited))
    return leaves


def build_dependency_graph(path):
    wb = load_workbook(path, data_only=False)
    graph = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                is_formula = isinstance(cell.value, str) and cell.value.startswith("=")

                fill = cell.fill
                rgb = None
                try:
                    rgb = fill.fgColor.rgb if fill and fill.fgColor else None
                except Exception:
                    rgb = None

                entry_key = f"{sheet}!{cell.coordinate}"
                direct_refs = []
                source_refs = []

                if is_formula:
                    direct_refs = [
                        f"{s}!{c}" for s, c in find_refs_in_formula(cell.value, sheet)
                    ]
                    try:
                        source_refs = resolve_to_source(wb, sheet, cell.coordinate)
                    except RecursionError:
                        source_refs = ["UNRESOLVED (max recursion depth exceeded)"]

                graph[entry_key] = {
                    "sheet": sheet,
                    "raw_value": cell.value if not is_formula else None,
                    "raw_formula": cell.value if is_formula else None,
                    "is_formula": is_formula,
                    "direct_refs": direct_refs,
                    "resolved_source_cells": source_refs,
                    "row_label": get_row_label(ws, cell.row),
                    "fill_rgb": rgb,
                    "is_green_input_candidate": is_greenish(rgb),
                }

    merged_ranges = {
        sheet: [str(r) for r in wb[sheet].merged_cells.ranges]
        for sheet in wb.sheetnames
    }

    return graph, merged_ranges


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_full_formula_chain.py <workbook.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    graph, merged_ranges = build_dependency_graph(path)

    output = {
        "dependency_graph": graph,
        "merged_cell_ranges": merged_ranges,
    }

    with open("formula_dependency_graph.json", "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(graph)} cell entries to formula_dependency_graph.json")


if __name__ == "__main__":
    main()
